import io
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse

router = APIRouter()


@router.post("/pdf-to-excel")
async def pdf_to_excel(
    files: list[UploadFile] = File(...),
    options: str = Form(default="{}"),
):
    if not files:
        raise HTTPException(status_code=400, detail="請上傳 PDF 檔案")

    try:
        import pdfplumber
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(status_code=501, detail="需要安裝 pdfplumber 和 openpyxl")

    data = await files[0].read()
    wb = Workbook()

    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for i, page in enumerate(pdf.pages):
            ws = wb.create_sheet(title=f"Page {i+1}")
            tables = page.extract_tables()
            if tables:
                row_idx = 1
                for table in tables:
                    for row in table:
                        for col_idx, cell in enumerate(row, 1):
                            ws.cell(row=row_idx, column=col_idx, value=cell or "")
                        row_idx += 1
                    row_idx += 1
            else:
                text = page.extract_text() or ""
                for row_idx, line in enumerate(text.split("\n"), 1):
                    ws.cell(row=row_idx, column=1, value=line)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=converted.xlsx"},
    )
